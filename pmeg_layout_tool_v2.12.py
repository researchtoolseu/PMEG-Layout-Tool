#!/usr/bin/env python3
"""
PMEG Fenestration Layout Tool – Version 2.12

Created by:
- Michael A. Lazaris
- Andreas M. Lazaris

© 2026

PMEG fenestration layout on millimeter paper (unrolled graft model).

Everything is referenced to the NOMINAL GRAFT DIAMETER (device), not the aorta.

Input formats:
- CSV format with metadata lines at top starting with '#', e.g.:
    # patient_name: John Doe
    # patient_age: 68
    # study_date: 2026-02-01
    # physician_name: Dr Jane Smith
    # graft_diam_mm: 30
    # paper: A4
    # orientation: portrait
    # film_height_mm: 90           (optional, default 90)
    # tie_num_rows: 3              (optional, default 3)
    # tie_edge_pad_mm: 8           (optional, default 8)
    # tie_positions_clock: 4,6,8   (optional, default 4,6,8)
    # cut_margin_mm: 20            (optional, default 20)

- NEW (anchors):
    # ap_anchor: SMA|CA|TOP|NONE       (default SMA)
    # v_anchor:  BELOW_RENALS|CA|SMA|TOP|NONE  (default BELOW_RENALS)

- Then a table with columns:
    name,y_mm,theta_deg,fen_diam_mm,notes

- OR (new, easier planning mode) you can omit y_mm and instead provide:
    name,theta_deg,fen_diam_mm,dist_from_zero_mm,notes

  with metadata:
    # proximal_cover_mm: 25
    # anchor: CA   (or SMA, etc.)

  Conventions for planning mode:
  - anchor = most proximal fenestration present (default: first row name if not given)
  - ZERO point = bottom edge of the anchor fenestration
  - dist_from_zero_mm = distance from ZERO (bottom of anchor fenestration) to the BOTTOM edge of each fenestration (mm).
  - The script converts these to y_mm (distance from graft TOP to fenestration CENTER) for plotting, derived from bottom-edge distances.


Conventions:
- y_mm: distance BELOW proximal fabric edge; y=0 is TOP; increasing y goes DOWN.
- theta_deg: signed; theta > 0 -> right of 0°, theta < 0 -> left of 0°.
- x_mm = (theta_deg / 360) * (π * D_graft)

Output:
- MAIN PDF + PNG (true scale) on A4/A3.
  Includes grid, labels, 100x100 mm calibration square, and wrap/cut guides.
- FILM PDF (true scale) for transparent film printing:
  No grid, minimal text, but INCLUDES 100x100 mm calibration square at the bottom of the page.
  Film uses full usable page height for trimming + scale verification; graft geometry is drawn within film_height_mm.

PRINTING (critical): Print PDFs at 100% / Actual Size. Disable Fit/Shrink/Scale-to-fit.
"""

from __future__ import annotations

import argparse
import datetime
import csv
import math
import re
import warnings
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import matplotlib
matplotlib.use("Agg")
import matplotlib.backends.backend_pdf  # ensure PyInstaller collects PDF backend
import matplotlib.pyplot as plt
from matplotlib.patches import Circle, Rectangle
from openpyxl import load_workbook

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
    module=r"openpyxl\.worksheet\._reader",
)

MM_PER_INCH = 25.4

PAPER_SIZES_MM = {
    "A4": (210.0, 297.0),
    "A3": (297.0, 420.0),
}

CUT_MARGIN_MM = 20.0  # default distance OUTSIDE graft edge used as cutting guide

# =========================
# AP marker controls
# =========================
AP_MARKER_LEN_MM = 8.0
AP_MARKER_COLOR = "red"
AP_MARKER_LW = 6.0
AP_MARKER_GAP_MM = 1.0          # gap below anchor circle before AP starts
AP_TOP_Y_MM = 10.0              # used when ap_anchor=TOP or anchor missing

# =========================
# CHECK marker controls (anti-180° rotation cue at 12 o'clock)
# - Draws a non-symmetric check mark "✓" to make 180° twist obvious.
# - Positioned by default BELOW renal fenestrations to avoid confusion with fenestrations.
# =========================
CHECK_MARKER_COLOR = AP_MARKER_COLOR
CHECK_MARKER_LW = None              # if None -> uses AP_MARKER_LW

# Geometry of "✓" marker:
#   left arm:  from (x0 - LEFT_W, yb - LEFT_H) to (x0, yb)
#   right arm: from (x0, yb) to (x0 + RIGHT_W, yb - RIGHT_H)
CHECK_LEFT_W_MM = 3.0
CHECK_LEFT_H_MM = 2.0
CHECK_RIGHT_W_MM = 8.0
CHECK_RIGHT_H_MM = 8.0

CHECK_ANCHOR_GAP_MM = 6.0           # vertical gap BELOW the anchor (e.g., below renal fenestrations)

# If renal anchors are missing, we fall back to a fixed y (TOP) or below AP start
CHECK_DEFAULT_BOTTOM_Y_MM = 30.0
CHECK_FALLBACK_BELOW_AP_MM = 8.0

# Optional: dot at the inferior point of CA on the 12 o'clock (AP) line
CA_BOTTOM_DOT_SIZE = 28
CA_BOTTOM_DOT_COLOR = AP_MARKER_COLOR

# =========================
# Dimension annotations (MAIN PDF only)
# =========================
DIM_LINE_LW = 1.2
DIM_TICK_LEN_MM = 3.5
DIM_OFFSET_FROM_WRAP_MM = 10.0   # distance to the right of the wrap edge (+half_perimeter)
DIM_LABEL_FONTSIZE = 8
DIM_LABEL_PAD_MM = 1.5

# A second, independent longitudinal dimension scale (edge-to-edge clearances)
DIM2_OFFSET_FROM_WRAP_MM = 24.0  # placed further right than DIM_OFFSET_FROM_WRAP_MM
DIM2_LABEL_FONTSIZE = 8
DIM2_LABEL_PAD_MM = 1.5

# =========================
# Reduction tie guide controls (styling)

# Reduction tie guide controls (styling) (styling)
# =========================
TIE_GUIDE_COLOR = "0.75"         # very light gray
TIE_GUIDE_LW = 1.0
TIE_GUIDE_STYLE = (0, (2, 3))    # dotted-like dash pattern
TIE_DOT_SIZE = 52
TIE_DOT_EDGE = "0.55"

# Defaults (can be overridden by CSV metadata)
DEFAULT_TIE_NUM_ROWS = 3
DEFAULT_TIE_EDGE_PAD_MM = 8.0
DEFAULT_TIE_POSITIONS_CLOCK = [4, 6, 8]


# =========================
# Data model
# =========================
@dataclass
class Target:
    name: str
    y_mm: float
    theta_deg: float
    fen_diam_mm: float
    notes: str = ""
    dist_from_zero_mm: Optional[float] = None

    def x_mm_for_graft(self, graft_circumference_mm: float) -> float:
        return (self.theta_deg / 360.0) * graft_circumference_mm


# =========================
# Helpers
# =========================
def slugify(s: str, max_len: int = 80) -> str:
    s = s.strip()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^A-Za-z0-9_\-]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s[:max_len] if len(s) > max_len else s


def normalize_study_date_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return ""
    m = re.match(r"^(\d{4}-\d{2}-\d{2})(?:[ T].*)?$", s)
    if m:
        return m.group(1)
    return s


def build_patient_line(metadata: Dict[str, str]) -> str:
    parts = []
    if metadata.get("patient_name"):
        parts.append(f"Patient: {metadata['patient_name']}")
    if metadata.get("patient_age"):
        parts.append(f"Age: {metadata['patient_age']}")
    study_date = normalize_study_date_text(metadata.get("study_date"))
    if study_date:
        parts.append(f"Date: {study_date}")
    return " | ".join(parts)


def get_physician_name(metadata: Dict[str, str]) -> str:
    for key in ("physician_name", "planner_name", "measuring_physician", "measured_by"):
        raw = (metadata.get(key, "") or "").strip()
        if raw:
            return raw
    return ""


def build_measurement_line(metadata: Dict[str, str]) -> str:
    physician_name = get_physician_name(metadata)
    if not physician_name:
        return ""
    return f"Measurements by: {physician_name}"


def auto_out_prefix_from_metadata(csv_path: Path, metadata: Dict[str, str], out_arg: Optional[str]) -> Path:
    if out_arg and out_arg.strip():
        return Path(out_arg)

    base_dir = csv_path.parent
    patient = metadata.get("patient_name", "").strip()
    date = metadata.get("study_date", "").strip()

    if patient or date:
        bits = []
        if patient:
            bits.append(slugify(patient))
        if date:
            bits.append(slugify(date))
        bits.append("PMEG")
        filename = "_".join(bits)
    else:
        filename = slugify(csv_path.stem) + "_PMEG"

    return base_dir / filename



# =========================
# Patient output folder structure (v1.4.7)
# =========================
def get_patients_root_dir(csv_path: Path) -> Path:
    """Default root directory for patient outputs.

    Preference:
    1) When packaged (PyInstaller), a 'Patients' folder alongside the executable.
    2) When running as a normal script, a 'Patients' folder alongside this script.
    3) If neither path is available for any reason, fall back to the input file directory.

    This keeps user-generated output outside the packaged app's internal folders.
    """
    try:
        if getattr(sys, "frozen", False):
            exe_dir = Path(sys.executable).resolve().parent
            return exe_dir / "Patients"
        script_dir = Path(__file__).resolve().parent
        return script_dir / "Patients"
    except Exception:
        return csv_path.parent / "Patients"




def get_runtime_info_dirs(input_path: Path) -> List[Path]:
    """Directories where small runtime sidecar files should be written.

    We write the sidecar in both:
    1) the executable/script directory
    2) the input workbook/file directory

    This makes Excel/VBA integration more robust even if the workbook is moved
    or launched from a slightly different location.
    """
    dirs: List[Path] = []

    try:
        if getattr(sys, "frozen", False):
            dirs.append(Path(sys.executable).resolve().parent)
        else:
            dirs.append(Path(__file__).resolve().parent)
    except Exception:
        pass

    try:
        dirs.append(input_path.resolve().parent)
    except Exception:
        dirs.append(input_path.parent)

    # remove duplicates while preserving order
    unique_dirs: List[Path] = []
    seen = set()
    for d in dirs:
        try:
            key = str(d.resolve())
        except Exception:
            key = str(d)
        if key not in seen:
            seen.add(key)
            unique_dirs.append(d)

    return unique_dirs


def write_last_output_folder_sidecar(input_path: Path, folder_path: Path) -> Optional[Path]:
    """Write the most recently created output folder path to one or more text files.

    The marker file is written both next to the executable/script and next to
    the input workbook/file, so that Excel/VBA can reliably find it.
    """
    written_path: Optional[Path] = None
    for sidecar_dir in get_runtime_info_dirs(input_path):
        try:
            sidecar_path = sidecar_dir / "last_output_folder.txt"
            sidecar_path.write_text(str(folder_path.resolve()), encoding="utf-8")
            if written_path is None:
                written_path = sidecar_path
        except Exception:
            pass
    return written_path


def _patient_folder_name(metadata: Dict[str, str]) -> str:
    """Create a stable patient folder name (date + patient name)."""
    patient = (metadata.get("patient_name") or "").strip()
    # Use study_date if provided; else today's date (local machine time)
    date = (metadata.get("study_date") or "").strip()
    if not date:
        date = datetime.date.today().isoformat()

    bits = []
    if date:
        bits.append(slugify(date))
    if patient:
        bits.append(slugify(patient))
    else:
        bits.append("UNKNOWN_PATIENT")

    return "_".join(bits)[:120]


def _next_version_index(patient_dir: Path) -> int:
    """Return the next integer version index for this patient (v001, v002...)."""
    if not patient_dir.exists():
        return 1

    max_v = 0
    for p in patient_dir.iterdir():
        if not p.is_dir():
            continue
        m = re.match(r"^v(\d{3})_", p.name)
        if m:
            try:
                max_v = max(max_v, int(m.group(1)))
            except Exception:
                pass
    return max_v + 1


def prepare_patient_run_dir(csv_path: Path, metadata: Dict[str, str]) -> Tuple[Path, Path]:
    """Create (Patients/<patient>/<v###_timestamp>/) and return (run_dir, out_prefix)."""
    patients_root = get_patients_root_dir(csv_path)
    patient_dir = patients_root / _patient_folder_name(metadata)

    # versioning + timestamp (unique & sortable)
    v_idx = _next_version_index(patient_dir)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = patient_dir / f"v{v_idx:03d}_{ts}"
    run_dir.mkdir(parents=True, exist_ok=True)

    # Copy the input file for traceability (keeps an immutable record of what was used)
    try:
        dst_input = run_dir / csv_path.name
        if not dst_input.exists():
            if csv_path.suffix.lower() in {".xlsx", ".xlsm"}:
                dst_input.write_bytes(csv_path.read_bytes())
            else:
                dst_input.write_text(csv_path.read_text(encoding="utf-8-sig"), encoding="utf-8")
    except Exception:
        # Non-fatal: still proceed with output generation
        pass

    out_prefix = run_dir / "PMEG_LAYOUT"
    return run_dir, out_prefix


def auto_out_prefix_and_dir(csv_path: Path, metadata: Dict[str, str], out_arg: Optional[str]) -> Tuple[Path, Optional[Path]]:
    """Return (out_prefix, run_dir).

    - If --out is provided, we respect it and do NOT auto-create patient folders (run_dir=None).
    - If --out is empty, we create Patients/<patient>/<v###_timestamp>/ and place outputs inside it.
    """
    if out_arg and out_arg.strip():
        return Path(out_arg), None
    run_dir, out_prefix = prepare_patient_run_dir(csv_path, metadata)
    return out_prefix, run_dir


def get_required_float(metadata: Dict[str, str], key: str) -> float:
    raw = (metadata.get(key, "") or "").strip()
    if not raw:
        raise ValueError(f"Missing required metadata: #{key}: <value>")
    val = float(raw)
    if val <= 0:
        raise ValueError(f"Metadata #{key} must be > 0. Got: {val}")
    return val


def get_optional_float(metadata: Dict[str, str], key: str, default: float) -> float:
    raw = (metadata.get(key, "") or "").strip()
    if not raw:
        return default
    val = float(raw)
    if val <= 0:
        raise ValueError(f"Metadata #{key} must be > 0. Got: {val}")
    return val


def get_optional_int(metadata: Dict[str, str], key: str, default: int, min_value: int = 1, max_value: int = 20) -> int:
    raw = (metadata.get(key, "") or "").strip()
    if not raw:
        return default
    val = int(raw)
    if val < min_value or val > max_value:
        raise ValueError(f"Metadata #{key} must be in [{min_value}, {max_value}]. Got: {val}")
    return val


def get_optional_str(metadata: Dict[str, str], key: str, default: str) -> str:
    raw = (metadata.get(key, "") or "").strip()
    return raw if raw else default


def get_paper(metadata: Dict[str, str]) -> str:
    paper = (metadata.get("paper", "A3") or "A3").strip().upper()
    if paper not in PAPER_SIZES_MM:
        raise ValueError(f"Metadata #paper must be one of {sorted(PAPER_SIZES_MM)}. Got: {paper}")
    return paper


def get_orientation(metadata: Dict[str, str]) -> str:
    ori = (metadata.get("orientation", "portrait") or "portrait").strip().lower()
    if ori not in {"portrait", "landscape"}:
        raise ValueError("Metadata #orientation must be 'portrait' or 'landscape'.")
    return ori


def find_target_by_name(targets: List[Target], name: str) -> Optional[Target]:
    name = name.strip().lower()
    for t in targets:
        if t.name.strip().lower() == name:
            return t
    return None


def safety_checks(targets: List[Target], film_height_mm: float) -> None:
    too_low = [t for t in targets if t.y_mm > film_height_mm]
    if too_low:
        print("\nWARNING: Some targets have y_mm > film_height_mm and will NOT appear in the FILM graft area:")
        for t in too_low:
            print(f"  - {t.name}: y={t.y_mm:.1f} mm (film_height_mm={film_height_mm:.1f})")
        print("  Tip: increase '# film_height_mm:' or reduce targets' y_mm if appropriate.\n")


# =========================
# Input reading
# =========================


def _normalize_header(s: object) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    aliases = {
        "angle_deg": "theta_deg",
        "angle": "theta_deg",
        "theta": "theta_deg",
        "diameter_mm": "fen_diam_mm",
        "fen_diameter_mm": "fen_diam_mm",
        "fenestration_diameter_mm": "fen_diam_mm",
        "distance_from_zero_mm": "dist_from_zero_mm",
        "dist0_mm": "dist_from_zero_mm",
        "distance_mm": "dist_from_zero_mm",
        "absolute_y_mm": "y_mm",
        "vessel": "name",
        "physician": "physician_name",
        "physician_name": "physician_name",
        "planner_name": "physician_name",
        "measuring_physician": "physician_name",
        "measured_by": "physician_name",
        "measurements_by": "physician_name",
        "cut_margin": "cut_margin_mm",
        "cut_margin_mm": "cut_margin_mm",
    }
    return aliases.get(s, s)


def _cell_to_str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
    return str(v).strip()


def _parse_targets_and_metadata_from_rows(raw_rows: List[Dict[str, str]], metadata: Dict[str, str], source_label: str) -> Tuple[List[Target], Dict[str, str]]:
    if not raw_rows:
        raise ValueError(f"{source_label} contains no target rows.")

    # Required core columns
    first_row_keys = set(raw_rows[0].keys())
    required_core = {"name", "theta_deg", "fen_diam_mm"}
    missing_core = required_core - first_row_keys
    if missing_core:
        raise ValueError(
            f"{source_label} table must contain columns: {sorted(required_core)} "
            f"(missing {sorted(missing_core)})"
        )

    # Optional / mode columns
    has_y = "y_mm" in first_row_keys
    has_dist = "dist_from_zero_mm" in first_row_keys

    if not has_y and not has_dist:
        raise ValueError(
            f"{source_label} must include either 'y_mm' (legacy mode) or 'dist_from_zero_mm' (planning mode)."
        )

    # Decide mode:
    planning_mode_needed = False
    if has_y:
        for row in raw_rows:
            if (row.get("y_mm") or "").strip() == "":
                planning_mode_needed = True
                break
    else:
        planning_mode_needed = True

    anchor_name = (metadata.get("anchor") or "").strip()
    if planning_mode_needed:
        if not has_dist:
            raise ValueError(
                "Some rows have blank 'y_mm' but the input has no 'dist_from_zero_mm' column. "
                "Add 'dist_from_zero_mm' or fill 'y_mm' for all rows."
            )
        if "proximal_cover_mm" not in metadata:
            raise ValueError(
                "Planning mode requires metadata 'proximal_cover_mm' "
                "(distance from graft TOP edge to TOP of the anchor fenestration)."
            )

    if not anchor_name:
        anchor_name = (raw_rows[0].get("name") or "").strip()

    diam_by_name: Dict[str, float] = {}
    for row in raw_rows:
        name = (row.get("name") or "").strip()
        if not name:
            raise ValueError("Target row with empty 'name'.")
        try:
            diam = float(row["fen_diam_mm"])
        except Exception:
            raise ValueError(f"Invalid fen_diam_mm for target '{name}': {row.get('fen_diam_mm')}")
        diam_by_name[name] = diam

    if planning_mode_needed and anchor_name not in diam_by_name:
        raise ValueError(f"Anchor '{anchor_name}' not found among target names in the input file.")

    if planning_mode_needed:
        try:
            proximal_cover_mm = float(metadata["proximal_cover_mm"])
        except Exception:
            raise ValueError(f"Invalid proximal_cover_mm metadata value: {metadata.get('proximal_cover_mm')}")

        anchor_diam = diam_by_name[anchor_name]
        y_anchor_center = proximal_cover_mm + (anchor_diam / 2.0)
        y_zero = proximal_cover_mm + anchor_diam
        metadata['_planning_mode'] = '1'
        metadata['_anchor_name'] = anchor_name
        metadata['_y_zero_mm'] = f"{y_zero:.3f}"
    else:
        y_anchor_center = None
        y_zero = None

    targets: List[Target] = []
    for row in raw_rows:
        name = (row.get("name") or "").strip()
        dist_for_table: Optional[float] = None

        y_val: Optional[float] = None
        if has_y:
            y_str = (row.get("y_mm") or "").strip()
            if y_str != "":
                try:
                    y_val = abs(float(y_str))
                except Exception:
                    raise ValueError(f"Invalid y_mm for target '{name}': {row.get('y_mm')}")

        if y_val is None:
            if not planning_mode_needed:
                raise ValueError(f"Target '{name}' has no y_mm, but planning mode is not enabled.")

            if name == anchor_name:
                y_val = float(y_anchor_center)
                dist_for_table = None
            else:
                dist_str = (row.get("dist_from_zero_mm") or "").strip()
                if dist_str == "":
                    raise ValueError(
                        f"Target '{name}' has blank y_mm and blank dist_from_zero_mm. Provide one of them."
                    )
                try:
                    dist = float(dist_str)
                except Exception:
                    raise ValueError(f"Invalid dist_from_zero_mm for target '{name}': {row.get('dist_from_zero_mm')}")
                dist_for_table = dist
                y_val = float(y_zero) + dist - (float(row["fen_diam_mm"]) / 2.0)

        targets.append(
            Target(
                name=name,
                y_mm=y_val,
                theta_deg=float(row["theta_deg"]),
                fen_diam_mm=float(row["fen_diam_mm"]),
                notes=(row.get("notes") or "").strip(),
                dist_from_zero_mm=dist_for_table,
            )
        )

    if not targets:
        raise ValueError(f"{source_label} contains no target rows.")
    return targets, metadata


def read_targets_and_metadata_xlsx(xlsx_path: Path) -> Tuple[List[Target], Dict[str, str]]:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["PMEG_Input"] if "PMEG_Input" in wb.sheetnames else wb[wb.sheetnames[0]]

    metadata_keys = {
        "patient_name", "patient_age", "study_date", "physician_name", "graft_diam_mm", "paper", "orientation",
        "film_height_mm", "tie_num_rows", "tie_edge_pad_mm", "tie_positions_clock", "cut_margin_mm", "dim_header_rotation",
        "ap_anchor", "v_anchor", "proximal_cover_mm", "anchor",
    }

    metadata: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=2, values_only=True):
        key = _normalize_header(row[0])
        if key in metadata_keys:
            val = _cell_to_str(row[1])
            if val != "":
                metadata[key] = val

    header_row_idx = None
    header_map: Dict[str, int] = {}
    for r in range(1, min(ws.max_row, 80) + 1):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 12) + 1)]
        norm = [_normalize_header(v) for v in vals]
        if {"name", "theta_deg", "fen_diam_mm"}.issubset(set(norm)):
            header_row_idx = r
            header_map = {n: i for i, n in enumerate(norm) if n}
            break

    if header_row_idx is None:
        raise ValueError("Could not find the target table header row in the Excel file.")

    raw_rows: List[Dict[str, str]] = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_dict: Dict[str, str] = {}
        populated = False
        for header, idx in header_map.items():
            val = ws.cell(r, idx + 1).value
            sval = _cell_to_str(val)
            row_dict[header] = sval
            if sval != "":
                populated = True
        if not populated:
            break
        if (row_dict.get("name") or "").startswith("#"):
            continue
        raw_rows.append(row_dict)

    return _parse_targets_and_metadata_from_rows(raw_rows, metadata, "Excel input")


def read_targets_and_metadata_csv(csv_path: Path) -> Tuple[List[Target], Dict[str, str]]:
    """
    Reads a targets CSV plus optional metadata lines starting with '#'.

    Supported modes:
    1) Absolute y mode (legacy):
       Columns: name,y_mm,theta_deg,fen_diam_mm[,notes]
       where y_mm is the distance from graft TOP edge to fenestration CENTER.

    2) Planning mode (new):
       Columns: name,theta_deg,fen_diam_mm,dist_from_zero_mm[,notes]
       (y_mm can be blank)

       Metadata required for planning mode:
         # proximal_cover_mm: <float>   (distance from graft TOP edge down to the TOP of the anchor fenestration)
       Metadata optional:
         # anchor: <name>              (fenestration name used as the most proximal target; default = first row name)

       Conventions:
         - ZERO point is the BOTTOM edge of the anchor fenestration.
         - dist_from_zero_mm is measured from ZERO down to the CENTER of each fenestration.
         - The script converts these to y_mm (from graft TOP to CENTER) for plotting.
    """
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    lines = csv_path.read_text(encoding="utf-8-sig").splitlines()

    metadata: Dict[str, str] = {}
    data_lines: List[str] = []

    for raw in lines:
        line = raw.strip()
        if not line:
            continue
        if line.startswith("#"):
            if ":" in line:
                key, value = line[1:].split(":", 1)
                key_norm = _normalize_header(key)
                if key_norm and re.fullmatch(r"[a-z0-9_]+", key_norm):
                    metadata[key_norm] = value.strip()
        else:
            data_lines.append(raw)

    reader = csv.DictReader(data_lines)
    if not reader.fieldnames:
        raise ValueError("CSV contains no header row.")

    # Required core columns
    required_core = {"name", "theta_deg", "fen_diam_mm"}
    missing_core = required_core - set(reader.fieldnames)
    if missing_core:
        raise ValueError(
            f"CSV table must contain columns: {sorted(required_core)} "
            f"(missing {sorted(missing_core)})"
        )

    # Optional / mode columns
    has_y = "y_mm" in (reader.fieldnames or [])
    has_dist = "dist_from_zero_mm" in (reader.fieldnames or [])

    if not has_y and not has_dist:
        raise ValueError(
            "CSV must include either 'y_mm' (legacy mode) or 'dist_from_zero_mm' (planning mode)."
        )

    raw_rows: List[Dict[str, str]] = []
    for row in reader:
        # Skip completely empty rows (common if someone leaves trailing lines)
        if not any((v or "").strip() for v in row.values()):
            continue
        raw_rows.append(row)

    if not raw_rows:
        raise ValueError("CSV contains no target rows.")

    # Decide mode:
    # - If any row has a non-empty y_mm, we treat it as absolute-y for that row.
    # - If y_mm is empty for a row, we will compute from dist_from_zero_mm (planning mode).
    planning_mode_needed = False
    if has_y:
        for row in raw_rows:
            if (row.get("y_mm") or "").strip() == "":
                # if y missing, we need planning info
                planning_mode_needed = True
                break
    else:
        planning_mode_needed = True

    # If planning mode is needed, ensure we have what we need.
    anchor_name = (metadata.get("anchor") or "").strip()
    if planning_mode_needed:
        if not has_dist:
            raise ValueError(
                "Some rows have blank 'y_mm' but CSV has no 'dist_from_zero_mm' column. "
                "Add 'dist_from_zero_mm' or fill 'y_mm' for all rows."
            )
        if "proximal_cover_mm" not in metadata:
            raise ValueError(
                "Planning mode requires metadata '# proximal_cover_mm: <mm>' "
                "(distance from graft TOP edge to TOP of the anchor fenestration)."
            )

    # Establish default anchor if not provided
    if not anchor_name:
        anchor_name = (raw_rows[0].get("name") or "").strip()

    # First, parse diameters (needed for anchor geometry)
    diam_by_name: Dict[str, float] = {}
    for row in raw_rows:
        name = (row.get("name") or "").strip()
        if not name:
            raise ValueError("Target row with empty 'name'.")
        try:
            diam = float(row["fen_diam_mm"])
        except Exception:
            raise ValueError(f"Invalid fen_diam_mm for target '{name}': {row.get('fen_diam_mm')}")
        diam_by_name[name] = diam

    if planning_mode_needed and anchor_name not in diam_by_name:
        raise ValueError(f"Anchor '{anchor_name}' not found among target names in CSV.")

    # Compute anchor-based conversion constants
    if planning_mode_needed:
        try:
            proximal_cover_mm = float(metadata["proximal_cover_mm"])
        except Exception:
            raise ValueError(f"Invalid proximal_cover_mm metadata value: {metadata.get('proximal_cover_mm')}")

        anchor_diam = diam_by_name[anchor_name]
        # y to anchor CENTER from graft top:
        y_anchor_center = proximal_cover_mm + (anchor_diam / 2.0)
        # ZERO is bottom edge of anchor fenestration:
        y_zero = proximal_cover_mm + anchor_diam
        metadata['_planning_mode'] = '1'
        metadata['_anchor_name'] = anchor_name
        metadata['_y_zero_mm'] = f"{y_zero:.3f}"
    else:
        y_anchor_center = None
        y_zero = None

    targets: List[Target] = []
    for row in raw_rows:
        name = (row.get("name") or "").strip()

        dist_for_table: Optional[float] = None  # planning-mode input distance (ZERO -> vessel center)

        # y_mm (absolute) if present and non-empty
        y_val: Optional[float] = None
        if has_y:
            y_str = (row.get("y_mm") or "").strip()
            if y_str != "":
                try:
                    y_val = abs(float(y_str))
                except Exception:
                    raise ValueError(f"Invalid y_mm for target '{name}': {row.get('y_mm')}")

        if y_val is None:
            # Compute from planning distances
            if not planning_mode_needed:
                raise ValueError(
                    f"Target '{name}' has no y_mm, but planning mode is not enabled."
                )

            if name == anchor_name:
                y_val = float(y_anchor_center)  # type: ignore[arg-type]
                dist_for_table = None  # ZERO is defined at the bottom of the anchor fenestration
            else:
                dist_str = (row.get("dist_from_zero_mm") or "").strip()
                if dist_str == "":
                    raise ValueError(
                        f"Target '{name}' has blank y_mm and blank dist_from_zero_mm. "
                        "Provide one of them."
                    )
                try:
                    dist = float(dist_str)
                except Exception:
                    raise ValueError(f"Invalid dist_from_zero_mm for target '{name}': {row.get('dist_from_zero_mm')}")
                dist_for_table = dist
                # dist is measured from ZERO (bottom of anchor) to the BOTTOM edge of this fenestration
                # Convert to fenestration CENTER for plotting
                y_val = float(y_zero) + dist - (float(row["fen_diam_mm"]) / 2.0)  # type: ignore[arg-type]

        targets.append(
            Target(
                name=name,
                y_mm=y_val,
                theta_deg=float(row["theta_deg"]),
                fen_diam_mm=float(row["fen_diam_mm"]),
                notes=(row.get("notes") or "").strip(),
                dist_from_zero_mm=dist_for_table,
            )
        )

    if not targets:
        raise ValueError("CSV contains no target rows.")
    return targets, metadata


def read_targets_and_metadata(input_path: Path) -> Tuple[List[Target], Dict[str, str]]:
    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        return read_targets_and_metadata_csv(input_path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_targets_and_metadata_xlsx(input_path)
    raise ValueError(f"Unsupported input file type: {input_path.suffix}. Use .csv or .xlsx")


# =========================
# Plot helpers
# =========================
def setup_mm_grid(ax, xlim, ylim):
    ax.set_xlim(*xlim)
    ax.set_ylim(*ylim)
    ax.invert_yaxis()

    ax.set_xticks(range(int(math.floor(xlim[0])), int(math.ceil(xlim[1])) + 1, 5))
    ax.set_yticks(range(int(math.floor(ylim[0])), int(math.ceil(ylim[1])) + 1, 5))
    ax.set_xticks(range(int(math.floor(xlim[0])), int(math.ceil(xlim[1])) + 1, 1), minor=True)
    ax.set_yticks(range(int(math.floor(ylim[0])), int(math.ceil(ylim[1])) + 1, 1), minor=True)

    ax.grid(which="major", linewidth=0.25)
    ax.grid(which="minor", linewidth=0.0)

    ax.set_aspect("equal", adjustable="box")
    ax.tick_params(
        axis="both",
        which="both",
        bottom=False,
        top=False,
        left=False,
        right=False,
        labelbottom=False,
        labelleft=False,
    )


def setup_film_axes(ax, xlim, ylim):
    ax.set_xlim(*xlim)
    ax.set_ylim(*ylim)
    ax.invert_yaxis()
    ax.set_aspect("equal", adjustable="box")
    ax.grid(False)
    ax.set_axis_off()


def draw_clockface(ax, graft_circumference_mm: float, y_offset=2.5, fontsize=7):
    clock_positions = [
        (0, "12"),
        (30, "1"), (60, "2"), (90, "3"), (120, "4"), (150, "5"),
        (180, "6"),
        (-150, "7"), (-120, "8"), (-90, "9"), (-60, "10"), (-30, "11"),
    ]
    half_perimeter = graft_circumference_mm / 2.0
    for angle, label in clock_positions:
        x = (angle / 360.0) * graft_circumference_mm
        if -half_perimeter <= x <= half_perimeter:
            ax.text(x, y_offset, label, ha="center", va="bottom", fontsize=fontsize, alpha=0.8)


def add_calibration_square(ax, xlim, ylim, size_mm: float = 100.0, pad_mm: float = 5.0, fontsize: int = 8):
    x0 = xlim[0] + pad_mm
    y0 = ylim[1] - pad_mm - size_mm
    ax.add_patch(Rectangle((x0, y0), size_mm, size_mm, fill=False, linewidth=2.0))
    ax.text(x0 + 2, y0 + 8, f"{int(size_mm)} mm", fontsize=fontsize, va="top")
def add_fenestration_table(
    ax,
    targets: List[Target],
    metadata: Dict[str, str],
    xlim,
    ylim,
    cal_square_mm: float = 100.0,
    pad_mm: float = 5.0,
):
    """Render a compact 'table-like' list of fenestrations on the left, above the calibration square.

    The on-plot labels are kept minimal (CA/SMA/RRA/LRA next to fenestrations), while
    all details are shown here to avoid collisions with markers/dimensions.
    """
    if not targets:
        return

    # Align table LEFT edge with the calibration square LEFT edge.
    square_x0 = xlim[0] + pad_mm
    x_box = square_x0

    # Place above the calibration square (bottom-left).
    y_square_top = ylim[1] - pad_mm - cal_square_mm
    y_top = max(0.0, y_square_top - 3.0)  # small gap above the square

    # Determine ZERO definition (planning mode)
    anchor_name = (metadata.get("_anchor_name") or metadata.get("anchor") or "").strip()
    planning_mode = (metadata.get("_planning_mode") == "1") or ("proximal_cover_mm" in metadata)
    if planning_mode and not anchor_name:
        # fallback: first row name
        anchor_name = targets[0].name

    zero_line = None
    if planning_mode and anchor_name:
        zero_line = f"Zero: bottom of {anchor_name} fen"

    # Summary metrics for the box
    ts_sorted = sorted(targets, key=lambda t: t.y_mm)
    r_prox = max(0.0, ts_sorted[0].fen_diam_mm / 2.0)
    proximal_len_above_top_fen_mm = max(0.0, ts_sorted[0].y_mm - r_prox)

    graft_diam_table = None
    try:
        graft_diam_table = float((metadata.get("graft_diam_mm") or metadata.get("graft_diameter_mm") or "").strip())
    except Exception:
        graft_diam_table = None

    # Build monospaced lines (ordered per your spec)
    lines: List[str] = []
    lines.append("Fenestrations")  # Line 1

    physician_line = build_measurement_line(metadata)
    if physician_line:
        lines.append(physician_line)

    if zero_line:
        lines.append(zero_line)   # Line 2

    # Lines 3-4: summary lines
    lines.append(f"Prox. length above top fen: {proximal_len_above_top_fen_mm:.0f} mm")
    if graft_diam_table is not None:
        lines.append(f"Graft Ø used for calc: {graft_diam_table:.1f} mm")
    else:
        lines.append("Graft Ø used for calc: —")

    lines.append("")  # empty spacer
    lines.append("Name   θ°   Ø mm   dist0   y(mm)")  # headings

    # Data rows
    for t in targets:
        dist0 = "—"
        if t.dist_from_zero_mm is not None:
            dist0 = f"{t.dist_from_zero_mm:>5.0f}"
        elif planning_mode and anchor_name and t.name.strip().lower() == anchor_name.strip().lower():
            dist0 = f"{0:>5.0f}"

        lines.append(
            f"{t.name:<5} {t.theta_deg:>4.0f} {t.fen_diam_mm:>5.0f} {dist0:>6} {t.y_mm:>7.1f}"
        )

    # Spacer + definitions (legend) under the table data
    lines.append("")
    lines.append("Definitions:")
    lines.append("name = target vessel")
    lines.append("θ° = clock-position angle of target vessel (degrees)")
    lines.append("Ø mm = fenestration diameter (mm)")
    lines.append("dist0 = bottom(anchor) → bottom(fenestration) (mm)")
    lines.append("y(mm) = top(graft) → center(fenestration) (mm)")

    # Compose text inside the "Fenestrations" box
    text = "\n".join(lines)

    # Background box size (approx; good enough for stable visuals)
    box_w = 82.0
    box_h = 3.2 * len(lines) + 0.0  # approx mm

    # Ensure the box doesn't run into the plot top
    y_box = max(0.0, y_top - box_h)

    ax.add_patch(
        Rectangle(
            (x_box, y_box),
            box_w,
            box_h,
            facecolor="white",
            edgecolor="black",
            linewidth=0.8,
            alpha=0.88,
            zorder=4,
        )
    )

    ax.text(
        x_box + 2.0,
        y_box + 2.0,
        text,
        fontsize=7,
        family="monospace",
        va="top",
        ha="left",
        zorder=5,
    )


def anchor_y_below_fenestration(targets: List[Target], anchor_name: str, gap_mm: float) -> Optional[float]:
    t = find_target_by_name(targets, anchor_name)
    if t is None or t.fen_diam_mm <= 0:
        return None
    r = t.fen_diam_mm / 2.0
    return max(0.0, t.y_mm + r + gap_mm)


def anchor_y_bottom_of_circle(targets: List[Target], anchor_name: str, gap_mm: float) -> Optional[float]:
    t = find_target_by_name(targets, anchor_name)
    if t is None or t.fen_diam_mm <= 0:
        return None
    r = t.fen_diam_mm / 2.0
    return max(0.0, t.y_mm + r + gap_mm)


def ap_marker_y_start(targets: List[Target], ap_anchor: str) -> Optional[float]:
    ap_anchor = ap_anchor.strip()

    if ap_anchor.upper() == "NONE":
        return None
    if ap_anchor.upper() == "TOP":
        return max(0.0, AP_TOP_Y_MM)

    # Generic vessel support
    t = find_target_by_name(targets, ap_anchor)
    if t is not None:
        return anchor_y_below_fenestration(targets, ap_anchor, AP_MARKER_GAP_MM)

    # fallback to SMA
    y = anchor_y_below_fenestration(targets, "SMA", AP_MARKER_GAP_MM)
    return y if y is not None else max(0.0, AP_TOP_Y_MM)

    if ap_anchor == "CA":
        y = anchor_y_below_fenestration(targets, "CA", AP_MARKER_GAP_MM)
        return y if y is not None else max(0.0, AP_TOP_Y_MM)

    y = anchor_y_below_fenestration(targets, "SMA", AP_MARKER_GAP_MM)
    return y if y is not None else max(0.0, AP_TOP_Y_MM)


def _renal_bottom_y(targets: List[Target], gap_mm: float) -> Optional[float]:
    """Return a y (mm) safely BELOW the lowest renal fenestration (inferior edge), plus gap."""
    renal_candidates: List[Target] = []
    for t in targets:
        name_u = t.name.strip().upper()
        name_l = t.name.strip().lower()
        # common renal labels: RRA/LRA, Right RA/Left RA, Renal, etc.
        tokens = re.split(r"[^A-Za-z0-9]+", name_u)
        if ("RENAL" in name_u) or ("RRA" in tokens) or ("LRA" in tokens) or ("RRA" in name_u) or ("LRA" in name_u):
            renal_candidates.append(t)
            continue
        if ("RIGHT" in name_u and "RA" in tokens) or ("LEFT" in name_u and "RA" in tokens):
            renal_candidates.append(t)

    if not renal_candidates:
        return None

    bottoms = []
    for t in renal_candidates:
        if t.fen_diam_mm <= 0:
            continue
        bottoms.append(t.y_mm + (t.fen_diam_mm / 2.0))
    if not bottoms:
        return None
    return max(0.0, max(bottoms) + gap_mm)


def check_marker_y_bottom(targets: List[Target], anchor: str, fallback_ap_y0: float) -> Optional[float]:
    anchor = anchor.strip()

    if anchor.upper() == "NONE":
        return None

    if anchor.upper() == "TOP":
        return max(0.0, CHECK_DEFAULT_BOTTOM_Y_MM)

    if anchor.upper() in {"RENALS", "BELOW_RENALS"}:
        y = _renal_bottom_y(targets, CHECK_ANCHOR_GAP_MM)
        if y is not None:
            return y
        y_ca = anchor_y_bottom_of_circle(targets, "CA", CHECK_ANCHOR_GAP_MM)
        if y_ca is not None:
            return y_ca
        return max(0.0, min(CHECK_DEFAULT_BOTTOM_Y_MM, fallback_ap_y0 + CHECK_FALLBACK_BELOW_AP_MM))

    # Generic vessel support
    t = find_target_by_name(targets, anchor)
    if t is not None:
        return anchor_y_bottom_of_circle(targets, anchor, CHECK_ANCHOR_GAP_MM)

    # fallback to BELOW_RENALS
    y = _renal_bottom_y(targets, CHECK_ANCHOR_GAP_MM)
    if y is not None:
        return y
    return max(0.0, CHECK_DEFAULT_BOTTOM_Y_MM)

    if anchor in {"RENALS", "BELOW_RENALS"}:
        y = _renal_bottom_y(targets, CHECK_ANCHOR_GAP_MM)
        if y is not None:
            return y
        # fallback: below CA if present
        y_ca = anchor_y_bottom_of_circle(targets, "CA", CHECK_ANCHOR_GAP_MM)
        if y_ca is not None:
            return y_ca
        return max(0.0, min(CHECK_DEFAULT_BOTTOM_Y_MM, fallback_ap_y0 + CHECK_FALLBACK_BELOW_AP_MM))

    if anchor == "CA":
        y = anchor_y_bottom_of_circle(targets, "CA", CHECK_ANCHOR_GAP_MM)
        if y is not None:
            return y
        return max(0.0, min(CHECK_DEFAULT_BOTTOM_Y_MM, fallback_ap_y0 + CHECK_FALLBACK_BELOW_AP_MM))

    if anchor == "SMA":
        y = anchor_y_bottom_of_circle(targets, "SMA", CHECK_ANCHOR_GAP_MM)
        if y is not None:
            return y
        return max(0.0, CHECK_DEFAULT_BOTTOM_Y_MM)

    # default: BELOW_RENALS
    y = _renal_bottom_y(targets, CHECK_ANCHOR_GAP_MM)
    if y is not None:
        return y
    return max(0.0, CHECK_DEFAULT_BOTTOM_Y_MM)


# =========================
# Tie guides
# =========================
def clock_hour_to_theta_deg(clock_hour: int) -> float:
    mapping = {
        12: 0.0,
        1: 30.0, 2: 60.0, 3: 90.0, 4: 120.0, 5: 150.0,
        6: 180.0,
        7: -150.0, 8: -120.0, 9: -90.0, 10: -60.0, 11: -30.0,
    }
    if clock_hour not in mapping:
        raise ValueError(f"Clock hour must be an integer from 1 to 12. Got: {clock_hour}")
    return mapping[clock_hour]


def parse_tie_positions_clock(metadata: Dict[str, str]) -> List[int]:
    raw = (metadata.get("tie_positions_clock", "") or "").strip()
    if not raw:
        return list(DEFAULT_TIE_POSITIONS_CLOCK)

    parts = [p.strip() for p in raw.split(",") if p.strip()]
    if not parts:
        return list(DEFAULT_TIE_POSITIONS_CLOCK)

    values: List[int] = []
    for p in parts:
        try:
            v = int(p)
        except Exception:
            raise ValueError(
                "Metadata #tie_positions_clock must be a comma-separated list of clock hours "
                f"(e.g. 4,6,8 or 5,6,7). Got: {raw}"
            )
        if v < 1 or v > 12:
            raise ValueError(
                "Metadata #tie_positions_clock must contain only clock hours from 1 to 12. "
                f"Got: {raw}"
            )
        values.append(v)

    if len(values) != 3:
        raise ValueError(
            "Metadata #tie_positions_clock must contain exactly 3 clock positions. "
            f"Got: {raw}"
        )

    if len(set(values)) != 3:
        raise ValueError(
            "Metadata #tie_positions_clock must contain 3 distinct clock positions. "
            f"Got: {raw}"
        )

    return values


def tie_positions_to_x_coords(clock_positions: List[int], graft_circumference: float, half_perimeter: float) -> List[float]:
    x_coords: List[float] = []
    for clock_hour in clock_positions:
        theta_deg = clock_hour_to_theta_deg(clock_hour)
        if clock_hour == 6:
            x_coords.extend([-half_perimeter, +half_perimeter])
        else:
            x_coords.append((theta_deg / 360.0) * graft_circumference)
    return x_coords


def tie_row_ys(y_min: float, y_max: float, n: int, edge_pad_mm: float) -> List[float]:
    if n <= 1:
        return [0.5 * (y_min + y_max)]
    span = max(1e-6, y_max - y_min)

    pad = min(edge_pad_mm, 0.25 * span)
    a = y_min + pad
    b = y_max - pad

    if b <= a:
        a = y_min + 2.0
        b = y_max - 2.0

    step = (b - a) / (n - 1)
    return [a + i * step for i in range(n)]


def draw_reduction_tie_guides(
    ax,
    graft_circumference: float,
    half_perimeter: float,
    *,
    y_min: float,
    y_max: float,
    film_mode: bool,
    row_ys: List[float],
    tie_positions_clock: List[int],
):
    # Reduction-tie guide positions on the unrolled graft are defined by clock positions
    # supplied in metadata (default: 4, 6, 8).
    x_positions = tie_positions_to_x_coords(tie_positions_clock, graft_circumference, half_perimeter)

    z = 3 if film_mode else 1
    for x in x_positions:
        ax.plot([x, x], [y_min, y_max],
                color=TIE_GUIDE_COLOR, linewidth=TIE_GUIDE_LW,
                linestyle=TIE_GUIDE_STYLE, zorder=z)

    for y in row_ys:
        ax.scatter(x_positions, [y] * len(x_positions),
                   s=TIE_DOT_SIZE, color=TIE_GUIDE_COLOR,
                   edgecolors=TIE_DOT_EDGE, linewidths=0.6, zorder=7)


# =========================
# AP + V drawing
# =========================
def draw_ap_and_v(ax, targets: List[Target], half_perimeter: float, *, with_text: bool, ap_anchor: str, v_anchor: str):
    """Draw AP (12/6) marker + a non-symmetric check mark (✓) at 12 o'clock as anti-180° twist cue.

    Note: we keep the historical parameter name 'v_anchor' for backward compatibility.
    """
    y0 = ap_marker_y_start(targets, ap_anchor)
    fallback_ap_y0 = AP_TOP_Y_MM if y0 is None else y0

    # -------------------------
    # AP marker (12/6)
    # -------------------------
    if y0 is not None:
        y1 = y0 + AP_MARKER_LEN_MM

        ax.plot([0, 0], [y0, y1],
                color=AP_MARKER_COLOR, linewidth=AP_MARKER_LW,
                solid_capstyle="round", zorder=6)

        for x6 in (-half_perimeter, +half_perimeter):
            ax.plot([x6, x6], [y0, y1],
                    color=AP_MARKER_COLOR, linewidth=AP_MARKER_LW,
                    solid_capstyle="round", zorder=6)

        ax.scatter([0, 0], [y0, y1],
                   s=TIE_DOT_SIZE, color=AP_MARKER_COLOR,
                   edgecolors=TIE_DOT_EDGE, linewidths=0.6, zorder=7)

        for x6 in (-half_perimeter, +half_perimeter):
            ax.scatter([x6, x6], [y0, y1],
                       s=TIE_DOT_SIZE, color=AP_MARKER_COLOR,
                       edgecolors=TIE_DOT_EDGE, linewidths=0.6, zorder=7)

        if with_text:
            ax.text(0, y1 + 2, "AP (12/6)",
                    ha="center", va="bottom", fontsize=7,
                    color=AP_MARKER_COLOR, alpha=0.9)

    # -------------------------
    # Optional: CA inferior dot on 12 o'clock line
    # -------------------------
    ca = find_target_by_name(targets, "CA")
    if ca is not None and ca.fen_diam_mm > 0:
        y_ca_inf = ca.y_mm + (ca.fen_diam_mm / 2.0)
        ax.scatter([0.0], [y_ca_inf],
                   s=CA_BOTTOM_DOT_SIZE, color=CA_BOTTOM_DOT_COLOR,
                   edgecolors=TIE_DOT_EDGE, linewidths=0.6, zorder=9)

    # -------------------------
    # Check mark (✓) – anti-180° cue at 12 o'clock
    # -------------------------
    yb = check_marker_y_bottom(targets, v_anchor, fallback_ap_y0=fallback_ap_y0)
    if yb is None:
        return

    lw = AP_MARKER_LW if CHECK_MARKER_LW is None else CHECK_MARKER_LW

    # Points: left tip -> bottom vertex -> right tip
    x0 = 0.0
    xL, yL = x0 - CHECK_LEFT_W_MM, yb - CHECK_LEFT_H_MM
    xR, yR = x0 + CHECK_RIGHT_W_MM, yb - CHECK_RIGHT_H_MM

    ax.plot([xL, x0], [yL, yb], color=CHECK_MARKER_COLOR, linewidth=lw,
            solid_capstyle="round", zorder=8)
    ax.plot([x0, xR], [yb, yR], color=CHECK_MARKER_COLOR, linewidth=lw,
            solid_capstyle="round", zorder=8)

    ax.scatter([xL, x0, xR], [yL, yb, yR],
               s=TIE_DOT_SIZE, color=CHECK_MARKER_COLOR,
               edgecolors=TIE_DOT_EDGE, linewidths=0.6, zorder=9)


# =========================
# Main plotting
# =========================

def write_report(
    out_prefix: Path,
    targets: List[Target],
    metadata: Dict[str, str],
    graft_diam_mm: float,
    paper: str,
    orientation: str,
    film_height_mm: float,
) -> Tuple[Path, Path]:
    """Write a human-readable TXT report + machine-friendly CSV report of inputs/derived values."""
    graft_circumference = math.pi * graft_diam_mm

    # sort by longitudinal position (y)
    ts = sorted(targets, key=lambda t: t.y_mm)

    # successive distances
    deltas = []
    for a, b in zip(ts, ts[1:]):
        deltas.append((a.name, b.name, b.y_mm - a.y_mm))

    # successive bottom-to-bottom distances:
    bottoms = []
    for t in ts:
        r = max(0.0, t.fen_diam_mm / 2.0)
        bottoms.append(t.y_mm + r)

    bottom_deltas = []
    for (a, yb_a), (b, yb_b) in zip(zip(ts, bottoms), zip(ts[1:], bottoms[1:])):
        bottom_deltas.append((a.name, b.name, yb_b - yb_a))

    txt_path = out_prefix.parent / (out_prefix.name + "_REPORT.txt")
    csv_path = out_prefix.parent / (out_prefix.name + "_REPORT.csv")

    lines = []
    lines.append(f"PMEG Fenestration Layout – Report")
    lines.append(f"Generated: {datetime.datetime.now().isoformat(timespec='seconds')}")
    physician_line = build_measurement_line(metadata)
    if physician_line:
        lines.append(physician_line)
    lines.append("")
    lines.append("=== Metadata (from input file) ===")
    if metadata:
        for k in sorted(metadata.keys()):
            lines.append(f"{k}: {metadata[k]}")
    else:
        lines.append("(none)")
    lines.append("")
    lines.append("=== Graft / Page ===")
    lines.append(f"paper: {paper}")
    lines.append(f"orientation: {orientation}")
    lines.append(f"nominal graft diameter (mm): {graft_diam_mm:.2f}")
    lines.append(f"graft circumference C=π·D (mm): {graft_circumference:.2f}")
    lines.append(f"film_height_mm (mm): {film_height_mm:.2f}")
    tie_positions_clock = parse_tie_positions_clock(metadata)
    lines.append(f"tie_positions_clock: {','.join(str(v) for v in tie_positions_clock)}")
    cut_margin_mm = get_optional_float(metadata, "cut_margin_mm", default=CUT_MARGIN_MM)
    lines.append(f"cut_margin_mm (mm): {cut_margin_mm:.2f}")
    lines.append("")
    lines.append("=== Fenestrations (sorted by distance from top) ===")
    header = "name | y_mm | clock(theta_deg) | x_mm (unrolled) | fen_diam_mm | notes"
    lines.append(header)
    lines.append("-" * len(header))
    for t in ts:
        x = t.x_mm_for_graft(graft_circumference)
        lines.append(f"{t.name} | {t.y_mm:.2f} | {t.theta_deg:.2f} | {x:.2f} | {t.fen_diam_mm:.2f} | {t.notes or ''}")
    lines.append("")
    lines.append("=== Longitudinal distances between successive fenestrations (Δy) ===")
    if deltas:
        for n1, n2, dy in deltas:
            lines.append(f"{n1} -> {n2}: {dy:.2f} mm")
    else:
        lines.append("(not applicable: fewer than 2 fenestrations)")
    lines.append("")

    lines.append("=== Bottom-to-bottom longitudinal distances between successive fenestrations (Δbottom) ===")
    if bottom_deltas:
        for n1, n2, gap in bottom_deltas:
            lines.append(f"{n1} -> {n2}: {gap:.2f} mm")
    else:
        lines.append("(not applicable: fewer than 2 fenestrations)")
    lines.append("")
    lines.append("=== Bottom-to-bottom distances from anchor to each fenestration (anchor bottom -> target bottom) ===")
    anchor_name = (metadata.get("_anchor_name") or metadata.get("anchor") or "").strip()
    if not anchor_name and ts:
        anchor_name = ts[0].name

    anchor_t = find_target_by_name(ts, anchor_name) if anchor_name else None
    if anchor_t is None and ts:
        anchor_t = ts[0]
        anchor_name = anchor_t.name

    if anchor_t is None:
        lines.append("(not available: anchor not found)")
    else:
        anchor_bottom = anchor_t.y_mm + max(0.0, anchor_t.fen_diam_mm / 2.0)
        for t in ts:
            t_bottom = t.y_mm + max(0.0, t.fen_diam_mm / 2.0)
            lines.append(f"{anchor_name} -> {t.name}: {t_bottom - anchor_bottom:.2f} mm")
    lines.append("")


    txt_path.write_text("\n".join(lines), encoding="utf-8")

    # CSV report
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["section", "key", "value"])
        physician_name = get_physician_name(metadata)
        if physician_name:
            w.writerow(["meta", "physician_name", physician_name])
        w.writerow(["meta", "paper", paper])
        w.writerow(["meta", "orientation", orientation])
        w.writerow(["meta", "graft_diam_mm", f"{graft_diam_mm:.2f}"])
        w.writerow(["meta", "graft_circumference_mm", f"{graft_circumference:.2f}"])
        w.writerow(["meta", "film_height_mm", f"{film_height_mm:.2f}"])
        w.writerow(["meta", "tie_positions_clock", ",".join(str(v) for v in parse_tie_positions_clock(metadata))])
        w.writerow(["meta", "cut_margin_mm", f"{get_optional_float(metadata, 'cut_margin_mm', default=CUT_MARGIN_MM):.2f}"])
        for k in sorted(metadata.keys()):
            w.writerow(["metadata", k, metadata[k]])
        w.writerow([])
        w.writerow(["fenestrations", "name", "y_mm", "theta_deg", "x_mm", "fen_diam_mm", "notes"])
        for t in ts:
            x = t.x_mm_for_graft(graft_circumference)
            w.writerow(["fenestration", t.name, f"{t.y_mm:.2f}", f"{t.theta_deg:.2f}", f"{x:.2f}", f"{t.fen_diam_mm:.2f}", t.notes])
        w.writerow([])
        w.writerow(["deltas", "from", "to", "delta_y_mm"])
        for n1, n2, dy in deltas:
            w.writerow(["delta", n1, n2, f"{dy:.2f}"])

        w.writerow([])
        w.writerow(["bottom_to_bottom", "from", "to", "delta_bottom_mm", "definition"])
        for n1, n2, gap in bottom_deltas:
            w.writerow(["delta_bottom", n1, n2, f"{gap:.2f}", "(bottom of lower) - (bottom of upper)"])

        w.writerow([])
        w.writerow(["anchor_bottom_to_bottom", "anchor", "to", "delta_mm", "definition"])
        anchor_name = (metadata.get("_anchor_name") or metadata.get("anchor") or "").strip()
        if not anchor_name and ts:
            anchor_name = ts[0].name
        anchor_t = find_target_by_name(ts, anchor_name) if anchor_name else None
        if anchor_t is None and ts:
            anchor_t = ts[0]
            anchor_name = anchor_t.name
        if anchor_t is not None:
            anchor_bottom = anchor_t.y_mm + max(0.0, anchor_t.fen_diam_mm / 2.0)
            for t in ts:
                t_bottom = t.y_mm + max(0.0, t.fen_diam_mm / 2.0)
                w.writerow(["anchor_bottom", anchor_name, t.name, f"{t_bottom - anchor_bottom:.2f}", "(target bottom) - (anchor bottom)"])

    return txt_path, csv_path


def draw_successive_distance_markings(
    ax,
    targets: List[Target],
    x_dim: float,
    label_side: str = "right",
):
    """Draw dimension ticks and labels between successive fenestrations (by y) on the MAIN plot."""
    ts = sorted(targets, key=lambda t: t.y_mm)
    if len(ts) < 2:
        return

    # ticks at each fenestration y
    for t in ts:
        ax.plot([x_dim, x_dim + DIM_TICK_LEN_MM], [t.y_mm, t.y_mm], linewidth=DIM_LINE_LW)

    # dimension segments + labels between successive y's
    for a, b in zip(ts, ts[1:]):
        y1, y2 = a.y_mm, b.y_mm
        ax.plot([x_dim, x_dim], [y1, y2], linewidth=DIM_LINE_LW)
        dy = y2 - y1
        ymid = (y1 + y2) / 2.0
        text = f"{dy:.0f} mm"
        if label_side == "right":
            ax.text(x_dim + DIM_TICK_LEN_MM + DIM_LABEL_PAD_MM, ymid, text,
                    ha="left", va="center", fontsize=DIM_LABEL_FONTSIZE,
                    bbox=dict(facecolor="white", edgecolor="none", pad=0.4, alpha=0.85))
        else:
            ax.text(x_dim - DIM_LABEL_PAD_MM, ymid, text,
                    ha="right", va="center", fontsize=DIM_LABEL_FONTSIZE,
                    bbox=dict(facecolor="white", edgecolor="none", pad=0.4, alpha=0.85))



def draw_successive_bottom_to_bottom_markings(
    ax,
    targets: List[Target],
    x_dim: float,
    label_side: str = "right",
):
    """Draw bottom-to-bottom longitudinal distances between successive fenestrations.

    Also includes a proximal segment from graft TOP (y=0) to the BOTTOM edge
    of the most proximal fenestration.
    """
    ts = sorted(targets, key=lambda t: t.y_mm)
    if not ts:
        return

    # bottom edges for each target
    bottoms = []
    for t in ts:
        r = max(0.0, t.fen_diam_mm / 2.0)
        bottoms.append(t.y_mm + r)

    # Proximal segment: TOP (0) -> bottom of most proximal fenestration
    yb0 = bottoms[0]
    ax.plot([x_dim, x_dim + DIM_TICK_LEN_MM], [0.0, 0.0], linewidth=DIM_LINE_LW)
    ax.plot([x_dim, x_dim + DIM_TICK_LEN_MM], [yb0, yb0], linewidth=DIM_LINE_LW)
    ax.plot([x_dim, x_dim], [0.0, yb0], linewidth=DIM_LINE_LW)

    ymid0 = (0.0 + yb0) / 2.0
    text0 = f"{yb0 - 0.0:.0f} mm"
    if label_side == "right":
        ax.text(
            x_dim + DIM_TICK_LEN_MM + DIM2_LABEL_PAD_MM,
            ymid0,
            text0,
            ha="left",
            va="center",
            fontsize=DIM2_LABEL_FONTSIZE,
            bbox=dict(facecolor="white", edgecolor="none", pad=0.4, alpha=0.85),
        )
    else:
        ax.text(
            x_dim - DIM2_LABEL_PAD_MM,
            ymid0,
            text0,
            ha="right",
            va="center",
            fontsize=DIM2_LABEL_FONTSIZE,
            bbox=dict(facecolor="white", edgecolor="none", pad=0.4, alpha=0.85),
        )

    if len(ts) < 2:
        return

    # For each pair, draw ticks at bottom edges, then a segment between them.
    for yb_a, yb_b in zip(bottoms, bottoms[1:]):
        ax.plot([x_dim, x_dim + DIM_TICK_LEN_MM], [yb_a, yb_a], linewidth=DIM_LINE_LW)
        ax.plot([x_dim, x_dim + DIM_TICK_LEN_MM], [yb_b, yb_b], linewidth=DIM_LINE_LW)

        ax.plot([x_dim, x_dim], [yb_a, yb_b], linewidth=DIM_LINE_LW)
        gap = yb_b - yb_a
        ymid = (yb_a + yb_b) / 2.0
        text = f"{gap:.0f} mm"
        if label_side == "right":
            ax.text(
                x_dim + DIM_TICK_LEN_MM + DIM2_LABEL_PAD_MM,
                ymid,
                text,
                ha="left",
                va="center",
                fontsize=DIM2_LABEL_FONTSIZE,
                bbox=dict(facecolor="white", edgecolor="none", pad=0.4, alpha=0.85),
            )
        else:
            ax.text(
                x_dim - DIM2_LABEL_PAD_MM,
                ymid,
                text,
                ha="right",
                va="center",
                fontsize=DIM2_LABEL_FONTSIZE,
                bbox=dict(facecolor="white", edgecolor="none", pad=0.4, alpha=0.85),
            )

def plot_layout_standard_paper(
    targets: List[Target],
    metadata: Dict[str, str],
    out_prefix: Path,
    base_title: str,
    patient_line: str,
    graft_diam_mm: float,
    paper: str,
    orientation: str,
    film_height_mm: float,
    tie_num_rows: int,
    tie_edge_pad_mm: float,
    tie_positions_clock: List[int],
    cut_margin_mm: float,
    margin_left_mm: float = 10.0,
    margin_right_mm: float = 10.0,
    margin_bottom_mm: float = 10.0,
    margin_top_mm: float = 25.0,
    title_fontsize: int = 11,
    patient_fontsize: int = 9,
    cal_square_mm: float = 100.0,
):
    pw, ph = PAPER_SIZES_MM[paper]
    if orientation == "landscape":
        pw, ph = ph, pw

    usable_w = pw - margin_left_mm - margin_right_mm
    usable_h = ph - margin_top_mm - margin_bottom_mm
    if usable_w <= 0 or usable_h <= 0:
        raise ValueError("Margins are too large for the selected paper size.")

    xlim = (-usable_w / 2.0, usable_w / 2.0)
    ylim = (0.0, usable_h)

    graft_circumference = math.pi * graft_diam_mm
    half_perimeter = graft_circumference / 2.0

    fig_w_in = pw / MM_PER_INCH
    fig_h_in = ph / MM_PER_INCH

    left = margin_left_mm / pw
    bottom = margin_bottom_mm / ph
    width = usable_w / pw
    height = usable_h / ph

    # MAIN
    fig = plt.figure(figsize=(fig_w_in, fig_h_in))
    ax = fig.add_axes([left, bottom, width, height])

    setup_mm_grid(ax, xlim, ylim)

    fig.suptitle(base_title, fontsize=title_fontsize, y=0.995)
    measurement_line = build_measurement_line(metadata)
    if patient_line.strip():
        fig.text(0.5, 0.971, patient_line, ha="center", va="top", fontsize=patient_fontsize)
    if measurement_line:
        measurement_y = 0.956 if patient_line.strip() else 0.971
        fig.text(0.5, measurement_y, measurement_line, ha="center", va="top", fontsize=8)

    if patient_line.strip() and measurement_line:
        graft_note_y = 0.940
    elif patient_line.strip() or measurement_line:
        graft_note_y = 0.948
    else:
        graft_note_y = 0.972
    fig.text(
        0.5,
        graft_note_y,
        f"Nominal graft Ø = {graft_diam_mm:.1f} mm  |  C = π·D = {graft_circumference:.1f} mm",
        ha="center",
        va="top",
        fontsize=8,
        alpha=0.9,
        bbox=dict(facecolor="white", edgecolor="none", pad=1.2),
    )

    add_calibration_square(ax, xlim, ylim, size_mm=cal_square_mm, pad_mm=5.0, fontsize=8)

    add_fenestration_table(ax, targets, metadata, xlim, ylim, cal_square_mm=cal_square_mm, pad_mm=5.0)

    ap_anchor = get_optional_str(metadata, "ap_anchor", "SMA")
    v_anchor = get_optional_str(metadata, "v_anchor", "BELOW_RENALS")
    draw_ap_and_v(ax, targets, half_perimeter, with_text=True, ap_anchor=ap_anchor, v_anchor=v_anchor)

    y_tie_max_main = min(film_height_mm, ylim[1])
    rows_main = tie_row_ys(0.0, y_tie_max_main, tie_num_rows, tie_edge_pad_mm)
    draw_reduction_tie_guides(
        ax, graft_circumference, half_perimeter,
        y_min=0.0, y_max=ylim[1], film_mode=False, row_ys=rows_main,
        tie_positions_clock=tie_positions_clock
    )

    ax.axvline(0, linewidth=2.0)
    ax.axvline(-half_perimeter, linewidth=2.0)
    ax.axvline(+half_perimeter, linewidth=2.0)

    ax.text(-half_perimeter, 3, "WRAP EDGE", ha="right", va="bottom",
            fontsize=7, rotation=90, alpha=0.8)
    ax.text(+half_perimeter, 3, "WRAP EDGE", ha="left", va="bottom",
            fontsize=7, rotation=90, alpha=0.8)

    cut_left = -half_perimeter - cut_margin_mm
    cut_right = +half_perimeter + cut_margin_mm
    ax.axvline(cut_left, linestyle=":", linewidth=1.2, alpha=0.8)
    ax.axvline(cut_right, linestyle=":", linewidth=1.2, alpha=0.8)

    ax.text(cut_left, 3, "CUT", ha="right", va="bottom",
            fontsize=7, rotation=90, alpha=0.6)
    ax.text(cut_right, 3, "CUT", ha="left", va="bottom",
            fontsize=7, rotation=90, alpha=0.6)

    draw_clockface(ax, graft_circumference_mm=graft_circumference, y_offset=2.5, fontsize=7)

    for t in targets:
        ax.plot([-half_perimeter, half_perimeter], [t.y_mm, t.y_mm], linewidth=1.0)

    for t in targets:
        x = t.x_mm_for_graft(graft_circumference)

        ax.add_patch(Circle((x, t.y_mm), t.fen_diam_mm / 2.0, fill=False, linewidth=2.0, zorder=5))

        font_size = max(8, t.fen_diam_mm * 1.2)
        ax.text(
            x, t.y_mm,
            f"{int(round(t.fen_diam_mm))}",
            ha="center", va="center",
            fontsize=font_size, fontweight="bold",
            zorder=6,
            bbox=dict(facecolor="white", edgecolor="none", pad=0.6),
        )
        # On-plot text labels moved to left-side table to avoid overlap.
        # Keep a minimal vessel label next to each fenestration (e.g., CA, SMA, RRA, LRA)
        ax.text(
            x + (t.fen_diam_mm / 2.0) + 2.0,
            t.y_mm,
            t.name,
            ha="left", va="center",
            fontsize=8,
            zorder=6,
            bbox=dict(facecolor="white", edgecolor="none", pad=0.2, alpha=0.75),
        )

        # (See add_fenestration_table.)

    # Dimension annotations: distances between successive fenestrations (MAIN PDF only)
    x_dim = +half_perimeter + DIM_OFFSET_FROM_WRAP_MM
    # NOTE: dimension scales live OUTSIDE the graft outline; they may extend into the CUT-MARGIN area.
    # Therefore we only constrain by the page x-limits (not by cut_right).
    if x_dim < (xlim[1] - 5.0):
        draw_successive_distance_markings(ax, targets, x_dim=x_dim, label_side="right")

        # A second (independent) vertical dimension scale:
        # edge-to-edge clearances (bottom of upper -> top of lower)
        x_dim2 = +half_perimeter + DIM2_OFFSET_FROM_WRAP_MM
        if x_dim2 < (xlim[1] - 5.0):
            draw_successive_bottom_to_bottom_markings(ax, targets, x_dim=x_dim2, label_side="right")
                        # Headers for the two scales (kept symmetric in y for clarity)
            y_header = 6.0
            rotation_angle = float(metadata.get("dim_header_rotation", "90") or "90")

            x_header_1 = x_dim + DIM_TICK_LEN_MM + 1.0
            x_header_2 = x_dim2 + DIM_TICK_LEN_MM + 1.0

            ax.text(
                x_header_1,
                y_header,
                "center-to-center",
                fontsize=7,
                fontweight="semibold",
                ha="left",
                va="bottom",
                rotation=rotation_angle,
                alpha=0.9,
                bbox=dict(facecolor="white", edgecolor="none", pad=0.25, alpha=0.85),
                zorder=10,
            )
            ax.text(
                x_header_2,
                y_header,
                "bottom-to-bottom",
                fontsize=7,
                fontweight="semibold",
                ha="left",
                va="bottom",
                rotation=rotation_angle,
                alpha=0.9,
                bbox=dict(facecolor="white", edgecolor="none", pad=0.25, alpha=0.85),
                zorder=10,
            )


    pdf_path = out_prefix.with_suffix(".pdf")
    png_path = out_prefix.with_suffix(".png")
    fig.savefig(pdf_path, format="pdf")
    fig.savefig(png_path, dpi=300)
    plt.close(fig)

    # FILM (full page usable height + bottom calibration square)
    film_prefix = out_prefix.parent / (out_prefix.name + "_FILM")
    film_pdf_path = film_prefix.with_suffix(".pdf")

    fig2 = plt.figure(figsize=(fig_w_in, fig_h_in))
    ax2 = fig2.add_axes([left, bottom, width, height])

    ylim_film_fullpage = (0.0, usable_h)
    setup_film_axes(ax2, xlim, ylim_film_fullpage)

    add_calibration_square(ax2, xlim, ylim_film_fullpage, size_mm=100.0, pad_mm=5.0, fontsize=8)

    add_fenestration_table(ax2, targets, metadata, xlim, ylim_film_fullpage, cal_square_mm=100.0, pad_mm=5.0)

    # graft area
    ax2.plot([cut_left, cut_right], [0.0, 0.0], linewidth=2.0)
    ax2.plot([0, 0], [0.0, film_height_mm], linestyle="--", linewidth=1.2)
    ax2.plot([cut_left, cut_right], [film_height_mm, film_height_mm], linewidth=2.0)

    ax2.plot([-half_perimeter, -half_perimeter], [0.0, film_height_mm], linewidth=2.0)
    ax2.plot([+half_perimeter, +half_perimeter], [0.0, film_height_mm], linewidth=2.0)

    ax2.plot([cut_left, cut_left], [0.0, film_height_mm], linestyle=":", linewidth=1.2)
    ax2.plot([cut_right, cut_right], [0.0, film_height_mm], linestyle=":", linewidth=1.2)

    draw_ap_and_v(ax2, targets, half_perimeter, with_text=False, ap_anchor=ap_anchor, v_anchor=v_anchor)

    rows_film = tie_row_ys(0.0, film_height_mm, tie_num_rows, tie_edge_pad_mm)
    draw_reduction_tie_guides(
        ax2, graft_circumference, half_perimeter,
        y_min=0.0, y_max=film_height_mm, film_mode=True, row_ys=rows_film,
        tie_positions_clock=tie_positions_clock
    )

    for t in targets:
        if t.y_mm > film_height_mm:
            continue
        x = t.x_mm_for_graft(graft_circumference)
        ax2.add_patch(Circle((x, t.y_mm), t.fen_diam_mm / 2.0, fill=False, linewidth=2.0, zorder=5))

        font_size = max(10, t.fen_diam_mm * 1.4)
        ax2.text(
            x, t.y_mm,
            f"{int(round(t.fen_diam_mm))}",
            ha="center", va="center",
            fontsize=font_size, fontweight="bold",
            zorder=6
        )

    R_X = +half_perimeter - 3.0
    R_Y = 6.0
    ax2.text(R_X, R_Y, "R", fontsize=18, fontweight="bold", ha="right", va="top")

    film_measurement_line = build_measurement_line(metadata)
    if film_measurement_line:
        fig2.text(0.99, 0.012, film_measurement_line, ha="right", va="bottom", fontsize=7, alpha=0.85)

    fig2.savefig(film_pdf_path, format="pdf")
    plt.close(fig2)

    # Reports (inputs + derived values)
    report_txt, report_csv = write_report(out_prefix, targets, metadata, graft_diam_mm, paper, orientation, film_height_mm)

    return pdf_path, png_path, film_pdf_path, report_txt, report_csv


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", dest="input_file", default="", help="Input CSV or XLSX file")
    ap.add_argument("--csv", dest="csv_legacy", default="", help=argparse.SUPPRESS)
    ap.add_argument("--out", default="", help="Output file prefix (optional). If empty, auto from metadata.")
    ap.add_argument("--title", default="PMEG Fenestration Layout", help="Top line title (line 1)")
    args = ap.parse_args()

    input_arg = (args.input_file or args.csv_legacy or "").strip()
    if not input_arg:
        ap.error("Provide --input <file.xlsx|file.csv> (or legacy --csv <file.csv>).")

    input_path = Path(input_arg)
    targets, metadata = read_targets_and_metadata(input_path)

    graft_diam_mm = get_required_float(metadata, "graft_diam_mm")
    paper = get_paper(metadata)
    orientation = get_orientation(metadata)
    film_height_mm = get_optional_float(metadata, "film_height_mm", default=90.0)

    tie_num_rows = get_optional_int(metadata, "tie_num_rows", default=DEFAULT_TIE_NUM_ROWS, min_value=1, max_value=10)
    tie_edge_pad_mm = get_optional_float(metadata, "tie_edge_pad_mm", default=DEFAULT_TIE_EDGE_PAD_MM)
    tie_positions_clock = parse_tie_positions_clock(metadata)
    cut_margin_mm = get_optional_float(metadata, "cut_margin_mm", default=CUT_MARGIN_MM)

    safety_checks(targets, film_height_mm)

    patient_line = build_patient_line(metadata)
    out_prefix, run_dir = auto_out_prefix_and_dir(input_path, metadata, args.out)

    pdf_path, png_path, film_pdf_path, report_txt, report_csv = plot_layout_standard_paper(
        targets=targets,
        metadata=metadata,
        out_prefix=out_prefix,
        base_title=args.title,
        patient_line=patient_line,
        graft_diam_mm=graft_diam_mm,
        paper=paper,
        orientation=orientation,
        film_height_mm=film_height_mm,
        tie_num_rows=tie_num_rows,
        tie_edge_pad_mm=tie_edge_pad_mm,
        tie_positions_clock=tie_positions_clock,
        cut_margin_mm=cut_margin_mm,
    )

    print(f"Saved (MAIN PDF):  {pdf_path}")
    print(f"Saved (PNG):       {png_path}")
    print(f"Saved (FILM PDF):  {film_pdf_path}")
    print(f"Saved (REPORT TXT): {report_txt}")
    print(f"Saved (REPORT CSV): {report_csv}")

    completed_folder = run_dir if run_dir is not None else out_prefix.parent
    sidecar_path = write_last_output_folder_sidecar(input_path, completed_folder)

    print(f"\nPatient output folder: {completed_folder}")
    if sidecar_path is not None:
        print(f"Last output folder pointer: {sidecar_path}")

    print("\nPRINTING (critical): Print PDFs at 100% / Actual Size. Disable Fit/Shrink/Scale-to-fit.")
    print("Check the 100 mm calibration square on BOTH the MAIN PDF and the FILM PDF with a ruler.")
    print("For transparent film, use the FILM PDF, verify the square, then trim along CUT lines.")
    print("\nPMEG Fenestration Layout Tool v2.12")


if __name__ == "__main__":
    main()
